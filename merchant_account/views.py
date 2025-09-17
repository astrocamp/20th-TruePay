from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth import login as django_login
from django.contrib.auth import logout as django_logout
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q, Avg, Case, When, IntegerField
from django.db.models.functions import TruncDate, ExtractHour
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


from truepay.decorators import no_cache_required, merchant_verified_required
from .forms import (
    RegisterForm,
    LoginForm,
    SubdomainChangeForm,
    MerchantProfileUpdateForm,
    MerchantOwnDomainForm,
)
from .services import DomainVerificationService
from customers_account.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from .models import Merchant, SubdomainRedirect, MerchantOwnDomain
from payments.models import Order, OrderItem, TicketValidation
from merchant_marketplace.models import Product
from payments.models import Order
from datetime import datetime


def register(req):
    if req.method == "POST":
        form = RegisterForm(req.POST)

        if form.is_valid():
            try:
                merchant = form.save()
                messages.success(req, "註冊成功！")
                return redirect("merchant_account:login")
            except Exception as e:
                messages.error(req, "註冊失敗，請重新再試")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(
                        req,
                        f"{form.fields[field].label if field in form.fields else field}: {error}",
                    )
    else:
        form = RegisterForm()

    return render(req, "merchant_account/Register.html", {"form": form})


def login(req):
    # 檢查用戶是否已經登入
    if req.user.is_authenticated and hasattr(req.user, 'member_type') and req.user.member_type == 'merchant':
        try:
            merchant = Merchant.objects.get(member=req.user)
            messages.info(req, "您已經登入了")
            return redirect("merchant_account:dashboard", merchant.subdomain)
        except Merchant.DoesNotExist:
            pass

    if req.method == "POST":
        form = LoginForm(req.POST)

        if form.is_valid():
            member = form.cleaned_data["member"]
            merchant = form.cleaned_data["merchant"]

            django_login(
                req, member, backend="django.contrib.auth.backends.ModelBackend"
            )

            # 檢查商家審核狀態
            if merchant.verification_status == "pending":
                messages.warning(req, "您的商家資料正在審核中，部分功能可能受限。審核通過後將開放完整功能。")
            elif merchant.verification_status == "rejected":
                messages.error(req, f"您的商家資料審核未通過。原因：{merchant.rejection_reason or '請聯絡客服了解詳情'}")
            elif merchant.verification_status == "suspended":
                messages.error(req, f"您的商家帳號已被暫停。原因：{merchant.rejection_reason or '請聯絡客服了解詳情'}")
            elif merchant.verification_status == "approved":
                messages.success(req, f"歡迎進入，{merchant.ShopName}！")
            else:
                messages.success(req, "歡迎進入！")

            return redirect("merchant_account:dashboard", merchant.subdomain)
        else:
            for error in form.non_field_errors():
                messages.error(req, error)
    else:
        form = LoginForm()

    return render(req, "merchant_account/login.html", {"form": form})


def logout(req):
    # 檢查用戶是否已經登入
    if not req.user.is_authenticated:
        messages.warning(req, "您尚未登入")
        return redirect("merchant_account:login")

    # 檢查是否為商家用戶
    if not hasattr(req.user, 'member_type') or req.user.member_type != 'merchant':
        messages.error(req, "權限不足")
        return redirect("merchant_account:login")

    django_logout(req)
    # 清除所有 session 資料
    req.session.flush()  # 完全清除 session 並重新生成 session key

    # 清除 messages（避免累積）
    storage = messages.get_messages(req)
    for message in storage:
        pass

    messages.success(req, "已成功登出")

    # 建立重導向回應並設定防快取 headers
    response = redirect("merchant_account:login")
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response


@no_cache_required
def dashboard(request, subdomain):
    """廠商Dashboard概覽頁面"""

    # 獲取商家的基本統計數據
    products = Product.objects.filter(merchant=request.merchant, is_active=True)
    total_products = products.count()
    recent_products = products.order_by("-created_at")[:5]

    # 獲取交易記錄統計
    orders = Order.objects.filter(product__merchant=request.merchant)
    total_orders = orders.count()
    recent_orders = orders.select_related("product", "customer").order_by(
        "-created_at"
    )[:5]

    # 計算總收入（只計算已付款的訂單，使用鏈式操作更高效）
    total_revenue = (
        orders.filter(status="paid").aggregate(total=Sum("amount"))["total"] or 0
    )

    # 票券統計數據
    tickets = OrderItem.objects.select_related("order", "product", "customer").filter(
        product__merchant=request.merchant
    )
    
    # 計算各種票券狀態
    now = timezone.now()
    ticket_stats = {
        'total': tickets.count(),
        'unused': tickets.filter(status='unused').count(),
        'used': tickets.filter(status='used').count(),
        'expired': tickets.filter(
            Q(status='expired') | 
            Q(valid_until__lt=now, status='unused')
        ).count()
    }
    
    # 計算即將到期的票券（未來24小時內）
    expiring_soon = tickets.filter(
        status='unused',
        valid_until__lte=now + timezone.timedelta(hours=24),
        valid_until__gt=now
    ).count()
    ticket_stats['expiring_soon'] = expiring_soon

    # 最近售出的票券（最近5張）
    recent_tickets = tickets.order_by('-created_at')[:5]

    context = {
        "merchant": request.merchant,
        "total_products": total_products,
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "recent_products": recent_products,
        "recent_orders": recent_orders,
        "ticket_stats": ticket_stats,
        "recent_tickets": recent_tickets,
    }

    return render(request, "merchant_account/dashboard.html", context)


@no_cache_required
def transaction_history(request, subdomain):
    """廠商交易記錄頁面"""
    # 查詢該商家的所有交易記錄（使用統一的 Order 模型）
    orders = (
        Order.objects.select_related("customer", "product")
        .prefetch_related("items")  # 預載入票券資料
        .filter(product__merchant=request.merchant)
        .order_by("-created_at")
    )

    order_items = orders

    # 分頁處理
    paginator = Paginator(order_items, 10)  # 每頁顯示10筆記錄
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "merchant": request.merchant,
        "page_obj": page_obj,
        "order_items": page_obj,
        "role": "merchant",  # 指定角色給模板使用
    }

    return render(request, "merchant_account/transaction_history.html", context)


# === 票券驗證相關視圖 ===


@no_cache_required
def ticket_validation_page(request, subdomain):
    """票券驗證主頁面"""
    return render(request, "merchant_account/ticket_validation.html")


@no_cache_required
@require_POST
def validate_ticket(request, subdomain):
    """驗證票券（手動輸入驗證碼或QR掃描）"""
    merchant = request.merchant
    raw_ticket_data = request.POST.get("ticket_code", "").strip()
    validation_method = request.POST.get("method", "manual")

    # 記錄驗證嘗試
    def create_validation_record(ticket, status, reason=""):
        if ticket:
            return TicketValidation.objects.create(
                ticket=ticket,
                merchant=merchant,
                status=status,
                failure_reason=reason,
                validation_method=validation_method,
                ip_address=request.META.get("REMOTE_ADDR"),
            )

    if not raw_ticket_data:
        context = {
            "error_message": "請輸入票券驗證碼",
            "merchant": merchant,
        }
        return render(request, "merchant_account/partials/ticket_error.html", context)

    # 處理QR code數據或手動輸入的票券代碼
    if validation_method == "qr":
        # QR code掃描：需要進行HMAC簽名驗證
        ticket, error_message = OrderItem.get_ticket_from_qr_data(raw_ticket_data)
        if not ticket:
            context = {
                "error_message": error_message or "QR code數據無效",
                "merchant": merchant,
            }
            return render(request, "merchant_account/partials/ticket_error.html", context)

        ticket_code = ticket.ticket_code
    else:
        # 手動輸入：直接查找票券代碼
        ticket_code = raw_ticket_data.upper()
        try:
            ticket = OrderItem.objects.select_related(
                "order", "product__merchant", "customer"
            ).get(ticket_code=ticket_code)
        except OrderItem.DoesNotExist:
            context = {
                "error_message": "找不到此票券代碼",
                "merchant": merchant,
            }
            return render(request, "merchant_account/partials/ticket_error.html", context)

    # 檢查票券有效性
    is_valid, message = ticket.is_valid()
    if not is_valid:
        create_validation_record(ticket, "failed", message)
        context = {
            "error_message": message,
            "merchant": merchant,
        }
        return render(
            request, "merchant_account/partials/ticket_error.html", context
        )

    # 檢查商家權限
    if ticket.product.merchant != merchant:
        create_validation_record(ticket, "unauthorized", "您無權限驗證此票券")
        context = {
            "error_message": "您無權限驗證此票券",
            "merchant": merchant,
        }
        return render(
            request, "merchant_account/partials/ticket_error.html", context
        )

    # 票券驗證成功，顯示確認頁面
    context = {
        "ticket_code": ticket_code,
        "ticket_info": ticket.ticket_info,
        "merchant": merchant,
    }
    return render(request, "merchant_account/partials/ticket_success.html", context)


@no_cache_required
@require_POST
def use_ticket(request, subdomain):
    """確認使用票券"""
    merchant = request.merchant
    ticket_code = request.POST.get("ticket_code", "").strip().upper()

    if not ticket_code:
        context = {
            "error_message": "票券代碼遺失",
            "merchant": merchant,
        }
        return render(request, "merchant_account/partials/ticket_error.html", context)

    try:
        ticket = OrderItem.objects.select_related(
            "order", "product__merchant", "customer"
        ).get(ticket_code=ticket_code)

        # 使用票券
        success, message = ticket.use_ticket(merchant)

        if success:
            # ticket.use_ticket() 已經更新票券狀態到資料庫，不需要額外記錄

            context = {
                "message": message,
                "ticket_value": ticket.order.unit_price,
                "used_at": ticket.used_at,
                "merchant": merchant,
            }
            return render(
                request, "merchant_account/partials/ticket_used.html", context
            )
        else:
            context = {
                "error_message": message,
                "merchant": merchant,
            }
            return render(
                request, "merchant_account/partials/ticket_error.html", context
            )

    except OrderItem.DoesNotExist:
        context = {
            "error_message": "找不到此票券代碼",
            "merchant": merchant,
        }
        return render(request, "merchant_account/partials/ticket_error.html", context)


@no_cache_required
@require_POST
def restart_scan(request, subdomain):
    """重新開始掃描"""
    context = {}
    return render(request, "merchant_account/partials/scan_ready.html", context)


@no_cache_required
def verification_records(request, subdomain):
    """票券使用紀錄頁面 - 顯示該商家的已使用票券記錄"""
    merchant = request.merchant

    # 取得篩選參數
    product_filter = request.GET.get("product", "")
    date_filter = request.GET.get("date", "")
    customer_filter = request.GET.get("customer", "")
    order_filter = request.GET.get("order", "")

    # 基本查詢：取得該商家的所有已使用票券
    used_tickets = (
        OrderItem.objects.select_related("order__customer__member", "product", "order")
        .filter(product__merchant=merchant, status="used")
        .order_by("-used_at")
    )

    # 商品篩選
    if product_filter:
        try:
            product_id = int(product_filter)
            used_tickets = used_tickets.filter(product_id=product_id)
        except (ValueError, TypeError):
            pass

    # 客戶篩選
    if customer_filter:
        used_tickets = used_tickets.filter(
            order__customer__member__email__icontains=customer_filter
        )

    # 日期篩選
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
            used_tickets = used_tickets.filter(used_at__date=filter_date)
        except ValueError:
            pass

    # 訂單編號篩選
    if order_filter:
        used_tickets = used_tickets.filter(
            order__provider_order_id__icontains=order_filter
        )

    # 統計資料（合併為單一 aggregate 查詢）

    all_used_tickets = OrderItem.objects.filter(
        product__merchant=merchant, status="used"
    )
    usage_stats = all_used_tickets.aggregate(
        total_tickets=Count("id"),
        total_revenue=Sum("order__unit_price"),
        today_tickets=Count("id", filter=Q(used_at__date=timezone.now().date())),
        products_count=Count("product", distinct=True),
    )
    usage_stats["total_revenue"] = usage_stats.get("total_revenue") or 0

    # 取得商品列表（用於篩選下拉選單）
    products = (
        Product.objects.filter(merchant=merchant, orderitem__status="used")
        .distinct()
        .order_by("name")
    )

    # 分頁處理
    paginator = Paginator(used_tickets, 15)  # 每頁顯示15筆記錄
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "merchant": merchant,
        "used_tickets": page_obj,
        "page_obj": page_obj,
        "usage_stats": usage_stats,
        "products": products,
        "product_filter": product_filter,
        "customer_filter": customer_filter,
        "date_filter": date_filter,
        "order_filter": order_filter,
    }

    return render(request, "merchant_account/verification_records.html", context)


@no_cache_required
def profile_settings(request, subdomain):
    """商家會員資料修改頁面"""
    merchant = request.merchant  # 由中間件提供

    if request.method == "POST":
        form_type = request.POST.get("form_type")

        if form_type == "profile":
            # 處理個人資料修改
            form = MerchantProfileUpdateForm(
                request.POST, instance=merchant, user=request.user
            )
            if form.is_valid():
                old_status = merchant.verification_status
                form.save()

                # 如果審核狀態改變了，顯示相應訊息
                merchant.refresh_from_db()  # 重新載入以獲取最新狀態
                if old_status != merchant.verification_status:
                    if merchant.verification_status == 'approved':
                        messages.success(request, "🎉 恭喜！您的商家資料已通過自動審核")
                    elif merchant.verification_status == 'rejected':
                        messages.warning(request, "商家資料已更新，但仍需完善部分資訊才能通過審核")
                else:
                    messages.success(request, "商家資料已成功更新")

                return redirect("merchant_account:profile_settings", subdomain)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(
                            request,
                            f"{form.fields[field].label if field in form.fields else field}: {error}",
                        )

        elif form_type == "password":
            # 處理密碼修改
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                password_form.save()
                # 更新 session，避免用戶被登出
                update_session_auth_hash(request, request.user)
                messages.success(request, "密碼已成功修改")
                return redirect("merchant_account:profile_settings", subdomain)
            else:
                for field, errors in password_form.errors.items():
                    for error in errors:
                        messages.error(
                            request,
                            f"{password_form.fields[field].label if field in password_form.fields else field}: {error}",
                        )

    # GET 請求或表單驗證失敗時顯示表單
    profile_form = MerchantProfileUpdateForm(instance=merchant, user=request.user)
    password_form = PasswordChangeForm(request.user)

    # 取得詳細的審核狀態資訊
    verification_info = merchant.get_verification_issues()

    context = {
        "merchant": merchant,
        "profile_form": profile_form,
        "password_form": password_form,
        "verification_info": verification_info,
    }

    return render(request, "merchant_account/profile_settings.html", context)


# 商家子網域
@no_cache_required
def subdomain_management(request, subdomain):
    merchant = request.merchant

    if request.method == "POST":
        form = SubdomainChangeForm(merchant, request.POST)
        if form.is_valid():
            try:
                new_subdomain = form.cleaned_data["new_subdomain"]
                reason = form.cleaned_data.get("reason", "商家主動修改")

                merchant.change_subdomain(new_subdomain, reason)

                messages.success(request, f"子網域已成功修改為 {new_subdomain}")

                return redirect("merchant_account:subdomain_management", new_subdomain)

            except ValueError as e:
                messages.error(request, str(e))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")

    can_change, status_message = merchant.can_change_subdomain()
    history = merchant.subdomain_history or []
    redirects = merchant.subdomain_redirects.filter(is_active=True)[:3]

    form = SubdomainChangeForm(merchant)

    context = {
        "merchant": merchant,
        "can_change": can_change,
        "status_message": status_message,
        "history": history[-3:],  # 三筆
        "redirects": redirects,
        "current_subdomain": merchant.subdomain,
        "form": form,
    }
    return render(request, "merchant_account/domain_settings.html", context)


@no_cache_required
def sold_tickets_list(request, subdomain):
    """廠商售出票券列表頁面"""
    merchant = request.merchant

    # 基本查詢：該廠商的所有已付款票券
    tickets = OrderItem.objects.filter(
        product__merchant=merchant,
        order__status='paid'
    ).select_related(
        'order', 'customer', 'customer__member', 'product'
    ).order_by('-created_at')

    # 篩選功能
    status_filter = request.GET.get('status', 'all')
    product_filter = request.GET.get('product', 'all')
    search_query = request.GET.get('search', '')

    # 狀態篩選
    if status_filter == 'unused':
        tickets = tickets.filter(status='unused')
    elif status_filter == 'used':
        tickets = tickets.filter(status='used')
    elif status_filter == 'expired':
        tickets = tickets.filter(status='expired')

    # 商品篩選
    if product_filter != 'all':
        try:
            product_id = int(product_filter)
            tickets = tickets.filter(product_id=product_id)
        except (ValueError, TypeError):
            pass

    # 搜尋功能（客戶姓名、客戶 email、訂單編號）
    if search_query:
        tickets = tickets.filter(
            Q(customer__name__icontains=search_query) |
            Q(customer__member__email__icontains=search_query) |
            Q(order__provider_order_id__icontains=search_query)
        )

    # 統計資訊
    all_tickets = OrderItem.objects.filter(
        product__merchant=merchant,
        order__status='paid'
    )

    stats = {
        'total_tickets': all_tickets.count(),
        'unused_tickets': all_tickets.filter(status='unused').count(),
        'used_tickets': all_tickets.filter(status='used').count(),
        'expired_tickets': all_tickets.filter(status='expired').count(),
        'total_revenue': all_tickets.aggregate(
            total=Sum('order__amount')
        )['total'] or 0,
    }

    # 即將到期的票券（7天內）
    from datetime import timedelta
    upcoming_expiry = timezone.now() + timedelta(days=7)
    stats['expiring_soon'] = all_tickets.filter(
        status='unused',
        valid_until__lte=upcoming_expiry,
        valid_until__gt=timezone.now()
    ).count()

    # 獲取廠商的所有商品（用於篩選下拉選單）
    merchant_products = Product.objects.filter(merchant=merchant).order_by('name')

    # 分頁功能
    paginator = Paginator(tickets, 20)  # 每頁20筆
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'merchant': merchant,
        'tickets': page_obj,
        'stats': stats,
        'merchant_products': merchant_products,
        'current_filters': {
            'status': status_filter,
            'product': product_filter,
            'search': search_query,
        },
        'status_choices': OrderItem.STATUS_CHOICES,
    }

    return render(request, 'merchant_account/sold_tickets_list.html', context)


# 商家自訂網域
@no_cache_required
def own_domain_list(request, subdomain):
    merchant = request.merchant
    domains = MerchantOwnDomain.objects.filter(merchant=merchant)
    form = MerchantOwnDomainForm()

    return render(
        request,
        "merchant_account/own_domain_list.html",
        {"merchant": merchant, "domains": domains, "form": form},
    )


@no_cache_required
def own_domain_add(request, subdomain):
    merchant = request.merchant

    if request.method == "POST":
        form = MerchantOwnDomainForm(request.POST)
        if form.is_valid():
            domain_obj = form.save(commit=False)
            domain_obj.merchant = merchant
            domain_obj.save()

            messages.success(
                request,
                f"網域 {domain_obj.domain_name} 已新增，請設定 DNS 記錄後進行驗證",
            )
            return redirect(
                "merchant_account:own_domain_detail",
                subdomain=subdomain,
                pk=domain_obj.pk,
            )
        else:
            for error in form.errors.values():
                messages.error(request, error[0])
    return redirect("merchant_account:own_domain_list", subdomain=subdomain)


@no_cache_required
def own_domain_detail(request, subdomain, pk):
    merchant = request.merchant
    merchant_domain = get_object_or_404(MerchantOwnDomain, pk=pk, merchant=merchant)
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "verify":
            success, message = DomainVerificationService.verify_domain_ownership(
                merchant_domain
            )
            if success:
                messages.success(request, message)
            else:
                messages.error(request, message)
        elif action == "delete":
            domain_name = merchant_domain.domain_name
            merchant_domain.delete()
            messages.success(request, f"網域 {domain_name} 已刪除")
            return redirect("merchant_account:own_domain_list", subdomain=subdomain)

    instructions = DomainVerificationService.get_verification_instructions(merchant_domain)
    return render(
        request,
        "merchant_account/own_domain_detail.html",
        {"merchant": merchant, "merchant_domain": merchant_domain, "instructions": instructions},
    )


# ===== 報表分析功能 =====


@no_cache_required
@merchant_verified_required
def reports_dashboard(request, subdomain):
    """報表分析總覽頁面"""
    merchant = request.merchant

    # 獲取時間範圍參數
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    # 基本統計數據
    orders = Order.objects.filter(product__merchant=merchant, created_at__gte=start_date)
    tickets = OrderItem.objects.filter(order__product__merchant=merchant, created_at__gte=start_date)
    products = Product.objects.filter(merchant=merchant)

    context = {
        'merchant': merchant,
        'days': days,
        'total_orders': orders.count(),
        'total_revenue': orders.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0,
        'total_tickets': tickets.count(),
        'total_products': products.count(),
        'ticket_usage_rate': round(tickets.filter(status='used').count() / max(tickets.count(), 1) * 100, 1),
    }

    return render(request, 'merchant_account/reports.html', context)


@no_cache_required
def export_sales_report(request, subdomain):
    """匯出銷售分析報表"""
    merchant = request.merchant
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    # 建立工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "銷售分析報表"

    # 設定樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # 報表標題
    ws['A1'] = f"{merchant.ShopName} - 銷售分析報表"
    ws['A2'] = f"統計期間：{start_date.date()} 至 {timezone.now().date()}"
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')

    # 設定標題樣式
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'].font = Font(size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')

    # 營收趨勢數據
    orders = Order.objects.filter(product__merchant=merchant, created_at__gte=start_date)

    # 總覽數據
    ws['A4'] = "營收總覽"
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill

    total_revenue = orders.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_orders = orders.count()
    paid_orders = orders.filter(status='paid').count()
    avg_order_value = orders.filter(status='paid').aggregate(Avg('amount'))['amount__avg'] or 0

    ws['A5'] = "總營收"
    ws['B5'] = f"NT$ {total_revenue:,}"
    ws['A6'] = "總訂單數"
    ws['B6'] = total_orders
    ws['A7'] = "成功訂單數"
    ws['B7'] = paid_orders
    ws['A8'] = "訂單成功率"
    ws['B8'] = f"{paid_orders/max(total_orders, 1)*100:.1f}%"
    ws['A9'] = "平均訂單金額"
    ws['B9'] = f"NT$ {avg_order_value:.0f}"

    # 訂單狀態分析
    ws['A11'] = "訂單狀態分析"
    ws['A11'].font = header_font
    ws['A11'].fill = header_fill

    status_stats = orders.values('status').annotate(count=Count('id')).order_by('-count')

    ws['A12'] = "狀態"
    ws['B12'] = "數量"
    ws['C12'] = "百分比"

    row = 13
    for stat in status_stats:
        status_display = dict(Order.STATUS_CHOICES).get(stat['status'], stat['status'])
        percentage = stat['count'] / max(total_orders, 1) * 100
        ws[f'A{row}'] = status_display
        ws[f'B{row}'] = stat['count']
        ws[f'C{row}'] = f"{percentage:.1f}%"
        row += 1

    # 金流方式統計
    ws[f"A{row + 1}"] = "金流方式統計"
    ws[f"A{row + 1}"].font = header_font
    ws[f"A{row + 1}"].fill = header_fill

    provider_stats = orders.values('provider').annotate(count=Count('id')).order_by('-count')

    row += 2
    ws[f'A{row}'] = "金流方式"
    ws[f'B{row}'] = "使用次數"
    ws[f'C{row}'] = "百分比"

    row += 1
    for stat in provider_stats:
        provider_display = dict(Order.PROVIDER_CHOICES).get(stat['provider'], stat['provider'])
        percentage = stat['count'] / max(total_orders, 1) * 100
        ws[f'A{row}'] = provider_display
        ws[f'B{row}'] = stat['count']
        ws[f'C{row}'] = f"{percentage:.1f}%"
        row += 1

    # 設定響應
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="銷售分析報表_{merchant.ShopName}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response


@no_cache_required
def export_ticket_report(request, subdomain):
    """匯出票券營運報表"""
    merchant = request.merchant
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    wb = Workbook()
    ws = wb.active
    ws.title = "票券營運報表"

    # 設定樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # 報表標題
    ws['A1'] = f"{merchant.ShopName} - 票券營運報表"
    ws['A2'] = f"統計期間：{start_date.date()} 至 {timezone.now().date()}"
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')

    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'].font = Font(size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')

    # 票券數據
    tickets = OrderItem.objects.filter(order__product__merchant=merchant, created_at__gte=start_date)

    # 票券總覽
    ws['A4'] = "票券營運總覽"
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill

    total_tickets = tickets.count()
    used_tickets = tickets.filter(status='used').count()
    unused_tickets = tickets.filter(status='unused').count()
    expired_tickets = tickets.filter(status='expired').count()
    usage_rate = used_tickets / max(total_tickets, 1) * 100

    ws['A5'] = "總票券數"
    ws['B5'] = total_tickets
    ws['A6'] = "已使用票券"
    ws['B6'] = used_tickets
    ws['A7'] = "未使用票券"
    ws['B7'] = unused_tickets
    ws['A8'] = "已過期票券"
    ws['B8'] = expired_tickets
    ws['A9'] = "票券使用率"
    ws['B9'] = f"{usage_rate:.1f}%"

    # 票券狀態分析
    ws['A11'] = "票券狀態分析"
    ws['A11'].font = header_font
    ws['A11'].fill = header_fill

    ws['A12'] = "狀態"
    ws['B12'] = "數量"
    ws['C12'] = "百分比"

    status_data = [
        ('未使用', unused_tickets),
        ('已使用', used_tickets),
        ('已過期', expired_tickets),
    ]

    row = 13
    for status, count in status_data:
        percentage = count / max(total_tickets, 1) * 100
        ws[f'A{row}'] = status
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{percentage:.1f}%"
        row += 1

    # 票券驗證統計
    validations = TicketValidation.objects.filter(
        ticket__order__product__merchant=merchant,
        validated_at__gte=start_date
    )

    ws[f'A{row + 1}'] = "票券驗證統計"
    ws[f'A{row + 1}'].font = header_font
    ws[f'A{row + 1}'].fill = header_fill

    row += 2
    total_validations = validations.count()
    successful_validations = validations.filter(status='success').count()
    success_rate = successful_validations / max(total_validations, 1) * 100

    ws[f'A{row}'] = "總驗證次數"
    ws[f'B{row}'] = total_validations
    row += 1
    ws[f'A{row}'] = "成功驗證次數"
    ws[f'B{row}'] = successful_validations
    row += 1
    ws[f'A{row}'] = "驗證成功率"
    ws[f'B{row}'] = f"{success_rate:.1f}%"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="票券營運報表_{merchant.ShopName}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response


@no_cache_required
def export_product_report(request, subdomain):
    """匯出商品表現報表"""
    merchant = request.merchant
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    wb = Workbook()
    ws = wb.active
    ws.title = "商品表現報表"

    # 設定樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # 報表標題
    ws['A1'] = f"{merchant.ShopName} - 商品表現報表"
    ws['A2'] = f"統計期間：{start_date.date()} 至 {timezone.now().date()}"
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')

    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'].font = Font(size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')

    # 商品銷售數據
    products = Product.objects.filter(merchant=merchant).annotate(
        order_count=Count('order', filter=Q(order__created_at__gte=start_date, order__status='paid')),
        revenue=Sum('order__amount', filter=Q(order__created_at__gte=start_date, order__status='paid'))
    ).order_by('-order_count')

    # 商品銷售排行
    ws['A4'] = "商品銷售排行榜"
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill

    # 設定表頭
    headers = ['排名', '商品名稱', '銷售數量', '總營收', '平均售價', '庫存數量', '狀態']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 填入商品數據
    row = 6
    total_revenue = 0
    for rank, product in enumerate(products, 1):
        revenue = product.revenue or 0
        total_revenue += revenue
        avg_price = revenue / product.order_count if product.order_count > 0 else product.price

        ws[f'A{row}'] = rank
        ws[f'B{row}'] = product.name
        ws[f'C{row}'] = product.order_count
        ws[f'D{row}'] = f"NT$ {revenue:,}"
        ws[f'E{row}'] = f"NT$ {avg_price:.0f}"
        ws[f'F{row}'] = product.stock
        ws[f'G{row}'] = "上架" if product.is_active else "下架"
        row += 1

    # 商品總結
    ws[f'A{row + 1}'] = "商品表現總結"
    ws[f'A{row + 1}'].font = header_font
    ws[f'A{row + 1}'].fill = header_fill

    row += 2
    active_products = products.filter(is_active=True).count()
    total_products = products.count()

    ws[f'A{row}'] = "商品總數"
    ws[f'B{row}'] = total_products
    row += 1
    ws[f'A{row}'] = "上架商品數"
    ws[f'B{row}'] = active_products
    row += 1
    ws[f'A{row}'] = "商品總營收"
    ws[f'B{row}'] = f"NT$ {total_revenue:,}"
    row += 1
    ws[f'A{row}'] = "平均商品營收"
    ws[f'B{row}'] = f"NT$ {total_revenue / max(total_products, 1):,.0f}"

    # 調整欄寬
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None

        # 找到第一個有效的儲存格來獲取欄位字母
        for cell in column_cells:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                break

        if column_letter is None:
            continue

        # 計算最大內容長度
        for cell in column_cells:
            if hasattr(cell, 'value') and cell.value is not None:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        # 設定欄寬
        adjusted_width = min(max(max_length + 2, 10), 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="商品表現報表_{merchant.ShopName}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response


# ===== 圖表數據API端點 =====

@no_cache_required
def get_sales_chart_data(request, subdomain):
    """獲取銷售分析圖表數據"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})

    try:
        merchant = request.merchant
        days = int(request.GET.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)

        # 營收趨勢數據 - 優化為單次查詢
        trend_data = []
        trend_labels = []

        # 使用 TruncDate 配合 annotate 一次性獲取所有日期的營收數據
        daily_revenues = Order.objects.filter(
            product__merchant=merchant,
            status='paid',
            created_at__gte=start_date
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            total_revenue=Sum('amount')
        ).order_by('date')

        revenue_dict = {item['date']: item['total_revenue'] for item in daily_revenues}

        # 生成完整的日期範圍（包含沒有資料的日期）
        current_date = start_date.date()
        end_date = timezone.now().date()

        while current_date <= end_date:
            daily_revenue = revenue_dict.get(current_date, 0)

            trend_labels.append(current_date.strftime('%m/%d'))
            trend_data.append(float(daily_revenue))
            current_date += timedelta(days=1)

        # 金流方式分析
        orders = Order.objects.filter(
            product__merchant=merchant,
            status='paid',
            created_at__gte=start_date
        )

        payment_stats = {}
        for order in orders:
            # 根據不同金流提供商確定支付方式
            if order.provider == 'newebpay':
                method = order.newebpay_payment_type or '藍新金流'
            elif order.provider == 'linepay':
                method = 'LINE Pay'
            else:
                method = order.get_provider_display() or '其他'

            if method not in payment_stats:
                payment_stats[method] = 0
            payment_stats[method] += float(order.amount)

        payment_labels = list(payment_stats.keys())
        payment_data = list(payment_stats.values())

        return JsonResponse({
            'success': True,
            'data': {
                'trend_labels': trend_labels,
                'trend_data': trend_data,
                'payment_labels': payment_labels,
                'payment_data': payment_data
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@no_cache_required
def get_tickets_chart_data(request, subdomain):
    """獲取票券營運圖表數據"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})

    try:
        merchant = request.merchant
        days = int(request.GET.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)

        tickets = OrderItem.objects.filter(
            order__product__merchant=merchant,
            created_at__gte=start_date
        )

        # 使用率統計
        used_tickets = tickets.filter(status='used').count()
        unused_tickets = tickets.filter(status='unused').count()
        expired_tickets = tickets.filter(status='expired').count()

        usage_labels = ['已使用', '未使用', '已過期']
        usage_data = [used_tickets, unused_tickets, expired_tickets]

        # 驗證時間分布（按時間段） - 優化為資料庫層面聚合查詢
        time_distribution = tickets.filter(
            status='used',
            used_at__isnull=False
        ).aggregate(
            morning=Count(Case(
                When(used_at__hour__gte=6, used_at__hour__lt=12, then=1),
                output_field=IntegerField()
            )),
            afternoon=Count(Case(
                When(used_at__hour__gte=12, used_at__hour__lt=18, then=1),
                output_field=IntegerField()
            )),
            evening=Count(Case(
                When(used_at__hour__gte=18, used_at__hour__lt=24, then=1),
                output_field=IntegerField()
            )),
            night=Count(Case(
                When(used_at__hour__gte=0, used_at__hour__lt=6, then=1),
                output_field=IntegerField()
            ))
        )

        time_stats = {
            '早上 (6-12)': time_distribution['morning'],
            '下午 (12-18)': time_distribution['afternoon'],
            '晚上 (18-24)': time_distribution['evening'],
            '深夜 (0-6)': time_distribution['night']
        }

        time_labels = list(time_stats.keys())
        time_data = list(time_stats.values())

        return JsonResponse({
            'success': True,
            'data': {
                'usage_labels': usage_labels,
                'usage_data': usage_data,
                'time_labels': time_labels,
                'time_data': time_data
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@no_cache_required
def get_products_chart_data(request, subdomain):
    """獲取商品表現圖表數據"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})

    try:
        merchant = request.merchant
        days = int(request.GET.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)

        # 商品銷售排行 (TOP 10) - 使用資料庫聚合查詢優化
        product_ranking = Order.objects.filter(
            product__merchant=merchant,
            status='paid',
            created_at__gte=start_date
        ).values('product__name').annotate(
            sales_count=Count('id')
        ).order_by('-sales_count')[:10]

        ranking_labels = [item['product__name'] for item in product_ranking]
        ranking_data = [item['sales_count'] for item in product_ranking]

        # 商品營收排行 (TOP 6) - 使用資料庫聚合查詢優化
        category_revenue = Order.objects.filter(
            product__merchant=merchant,
            status='paid',
            created_at__gte=start_date
        ).values('product__name').annotate(
            total_revenue=Sum('amount')
        ).order_by('-total_revenue')[:6]

        # 使用完整商品名稱作為標籤
        revenue_labels = [item['product__name'] for item in category_revenue]
        revenue_data = [float(item['total_revenue']) for item in category_revenue]

        return JsonResponse({
            'success': True,
            'data': {
                'ranking_labels': ranking_labels,
                'ranking_data': ranking_data,
                'revenue_labels': revenue_labels,
                'revenue_data': revenue_data
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
